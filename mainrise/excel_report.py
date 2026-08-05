"""跟踪报告 Excel 输出：多 Sheet 表格 + 表头样式 + A股红涨绿跌配色。"""
from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
RED = Font(color="C00000", bold=True)
GREEN = Font(color="008000", bold=True)
THIN = Border(*[Side(style="thin", color="D9D9D9")] * 4)


def _fmt(v, nd: int = 2) -> str:
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return "-"
    if isinstance(v, (int, float, np.floating, np.integer)):
        return round(float(v), nd)
    return str(v)


def _fill_sheet(ws, headers: list[str], rows: list[list], pct_cols: tuple = ()) -> None:
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN
    for row in rows:
        ws.append([_fmt(v) for v in row])
    for r in range(2, ws.max_row + 1):
        for c in pct_cols:
            v = ws.cell(r, c).value
            if isinstance(v, (int, float)):
                ws.cell(r, c).font = RED if v > 0 else (GREEN if v < 0 else Font())
        for c in range(1, len(headers) + 1):
            ws.cell(r, c).border = THIN
    for i, h in enumerate(headers, 1):
        widths = [len(str(ws.cell(r, i).value or "")) for r in range(1, ws.max_row + 1)]
        ws.column_dimensions[get_column_letter(i)].width = min(max(max(widths) + 2, 10), 60)
    ws.freeze_panes = "A2"


def write_tracking_excel(path: Path, date: str, buy_points: pd.DataFrame,
                         positions: pd.DataFrame, status_df: pd.DataFrame,
                         found: pd.DataFrame) -> Path:
    wb = Workbook()

    # Sheet1 今日买点提示
    ws = wb.active
    ws.title = "今日买点提示"
    rows = [[r["code"], r["name"], _fmt(r["composite"], 1), r["status"], r["hint"]]
            for _, r in buy_points.iterrows()] if not buy_points.empty else [["-", "-", "-", "无触发", "今日无买点"]]
    _fill_sheet(ws, ["代码", "名称", "综合分", "状态", "提示"], rows)

    # Sheet2 持仓管理（活跃+待买入）
    ws = wb.create_sheet("持仓管理")
    hold = positions[positions["status"].isin(["active", "pending"])] if not positions.empty \
        else pd.DataFrame()
    if not hold.empty and "pnl" in hold.columns:
        rows = [[r["code"], r["name"], r["buy_date"], _fmt(r["buy_price"]),
                 _fmt(r["last_close"]), _fmt(r["pnl"]), _fmt(r["peak"]),
                 r["status"], r["reason"] or ""] for _, r in hold.iterrows()]
        _fill_sheet(ws, ["代码", "名称", "买入日", "买入价", "现价", "盈亏%",
                         "峰值", "状态", "提示"], rows, pct_cols=(6,))
    else:
        _fill_sheet(ws, ["代码", "名称", "买入日", "买入价", "现价", "盈亏%",
                         "峰值", "状态", "提示"], [["-", "当前无持仓", "-", "-", "-", "-", "-", "-", "-"]])

    # Sheet3 历史平仓
    ws = wb.create_sheet("历史平仓")
    closed = positions[positions["status"] == "closed"] if not positions.empty \
        else pd.DataFrame()
    if closed.empty:
        _fill_sheet(ws, ["代码", "名称", "买入日", "买入价", "平仓日", "平仓价", "原因"],
                    [["-", "暂无平仓记录", "-", "-", "-", "-", "-"]])
    else:
        rows = [[r["code"], r["name"], r["buy_date"], _fmt(r["buy_price"]),
                 r["close_date"], _fmt(r["close_price"]), r["reason"]]
                for _, r in closed.iterrows()]
        _fill_sheet(ws, ["代码", "名称", "买入日", "买入价", "平仓日", "平仓价", "原因"], rows)

    # Sheet4 观察池状态
    ws = wb.create_sheet("观察池状态")
    rows = []
    for i, (_, r) in enumerate(status_df.iterrows(), 1):
        rows.append([i, r["code"], r["name"], _fmt(r["composite"], 1),
                     _fmt(r["close"]), _fmt(r["chg"]), r["status"], r["hint"],
                     _fmt(r["ma10"]), _fmt(r["ma20"]), _fmt(r["vr"], 2), _fmt(r["chg10"], 1)])
    _fill_sheet(ws, ["排名", "代码", "名称", "综合分", "收盘", "涨跌%", "状态", "提示",
                     "MA10", "MA20", "量比", "10日涨幅%"], rows, pct_cols=(6, 12))

    # Sheet5 全市场新信号
    ws = wb.create_sheet("全市场新信号")
    if found.empty:
        _fill_sheet(ws, ["代码", "名称", "类型", "涨幅%", "量比", "10日涨幅%", "说明"],
                    [["-", "当日无新信号", "-", "-", "-", "-", "-"]])
    else:
        rows = [[r["code"], r["name"], "T0信号日" if r["kind"] == "T0" else "T1确认日",
                 _fmt(r["chg"]), _fmt(r["vr"], 2), _fmt(r["chg10"], 1), "候选待评估"]
                for _, r in found.iterrows()]
        _fill_sheet(ws, ["代码", "名称", "类型", "涨幅%", "量比", "10日涨幅%", "说明"],
                    rows, pct_cols=(4, 6))

    # Sheet6 规则
    ws = wb.create_sheet("规则与说明")
    rules = [
        ("信号日 T0", "均线多头(MA5>MA10>MA20) + 收盘创20日新高 + (涨幅≥5%且量比≥1.5)或涨停"),
        ("买点1", "T0次日确认(收盘>MA5 且 低点≥T0收盘×0.97) → 次日开盘买入"),
        ("买点2", "确认后回踩 MA10 缩量企稳(低点≥MA20×0.99) → 低吸"),
        ("止损", "买入价 -5%，或跌破 MA10"),
        ("止盈", "高点回落 8%（条件单）"),
        ("时间止损", "持仓 5 个交易日"),
        ("仓位", "单票 ≤1/3，最多 3 只并行，优先综合分 Top10"),
        ("防追高", "10 个交易日涨幅 ≥80% 不进买点提示"),
        ("综合评分", "40%财务质量 + 30%信号强度 + 30%产业链地位"),
        ("数据截止", date),
        ("免责", "研究线索，不构成投资建议"),
    ]
    _fill_sheet(ws, ["项目", "说明"], [[k, v] for k, v in rules])
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 70

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path
