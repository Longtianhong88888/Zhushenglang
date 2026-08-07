"""KPI 仪表盘自动更新。

以 QuantDark 主题模板为底（优先 output/reports/主升浪跟踪仪表盘_QuantDark.xlsx，
打包软件用内置 mainrise/resources/dashboard_template.xlsx），读取 output/reports 下
全部跟踪 CSV（主升浪跟踪_*.csv）与持仓（output/state/mainrise_positions.csv），
重新生成最终仪表盘 output/reports/主升浪跟踪仪表盘.xlsx：

- Data:     全部交易日明细（观察池 = 有综合分的行，其余为新信号），"买点序号"列标记最新日买点行
- Watch:    最新日观察池按综合分降序 + >80(红)/75-80(橙) 分档列
- Summary:  按日聚合（天数增长时各区块自动下移）+ 状态分布 + 持仓收益（每日收盘/买入价-1）
- Dashboard: KPI 卡片、今日买点明细、三张图表（图表 XML 仅改数据引用，样式原样保留）

用法:
    python3 -m mainrise.cli dashboard          # 手动更新
    mainrise track                             # 跟踪报告生成后自动同步更新
"""
from __future__ import annotations

import csv
import re
import shutil
import sys
import xml.etree.ElementTree as ET
import zipfile
from copy import copy
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from mainrise import paths

RESOURCES = Path(__file__).resolve().parent / "resources"
BUNDLED_TEMPLATE = RESOURCES / "dashboard_template.xlsx"
QUANTDARK_TEMPLATE_NAME = "主升浪跟踪仪表盘_QuantDark.xlsx"
OUTPUT_NAME = "主升浪跟踪仪表盘.xlsx"
TRACK_PREFIX = "主升浪跟踪_"

BUY_STATUSES = ("T1确认买点", "T0新信号", "回踩低吸")
STATUSES = ("空头", "破位", "多头持有", "回踩低吸", "多头回踩", "T1确认买点", "T0新信号")
DEFAULT_MAX_10D_GAIN = 80.0

# 模板"原型"坐标：n=3 天时 Summary 布局与模板完全一致，n 增大时作为样式复制来源
PER_DAY_PROTO = {c: f"{c}2" for c in "ABCDEFGHIJKL"}
CROSS_SECTION_PROTO = "A5"
CROSS_HEADER_STATUS_PROTO = "A6"
CROSS_HEADER_DATE_PROTO = "B6"
CROSS_STATUS_LABEL_PROTO = "A7"
CROSS_STATUS_DATA_PROTO = "B7"
NOTE_PROTO = "A14"
POS_SECTION_PROTO = "A16"
POS_HEADER_PRICE_PROTO = "O17"
POS_HEADER_PNL_PROTO = "P17"
POS_DATA_PRICE_PROTO = "O18"
POS_DATA_PNL_PROTO = "P18"
BUY_LABEL_PROTO = "P22"
BUY_VALUE_PROTO = "P23"
BUY_NOTE_PROTO = "P24"

MAX_DATA_ROW = 5000  # Summary/Dashboard 公式引用的 Data 行上限（留足增长空间）


# ---------------------------------------------------------------- 数据读取

def _num(value: str) -> float | None:
    """CSV 数值转 float；空串/非法返回 None。"""
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _tracking_csvs(reports_dir: Path) -> list[tuple[date, Path]]:
    out = []
    for p in sorted(reports_dir.glob(f"{TRACK_PREFIX}*.csv")):
        stem = p.stem[len(TRACK_PREFIX):]  # 2026-08-06
        try:
            d = date.fromisoformat(stem)
        except ValueError:
            continue
        out.append((d, p))
    return out


def _parse_csv(path: Path) -> list[list[str]]:
    with open(path, encoding="utf-8-sig", newline="") as f:
        rows = list(csv.reader(f))
    return rows[1:] if rows else []


def _watch_count(raw: list[list[str]]) -> int:
    """观察池大小：CSV 开头连续有综合分的行数（新信号行综合分为空）。"""
    n = 0
    for r in raw:
        if len(r) > 2 and str(r[2]).strip():
            n += 1
        else:
            break
    return n


def _data_rows(csvs: list[tuple[date, Path]]) -> tuple[list[list], int, list[int], int]:
    """合并全部跟踪 CSV -> Data 明细行；返回 (rows, 最新日行数, 最新日买点行号, 观察池数量)。"""
    rows: list[list] = []
    latest = csvs[-1][0]
    latest_buy_rows: list[int] = []
    for d, p in csvs:
        raw = _parse_csv(p)
        wc = _watch_count(raw)
        for i, r in enumerate(raw):
            group = "观察池" if i < wc else "新信号"
            code, name = (r[0].strip(), r[1].strip()) if len(r) > 1 else ("", "")
            row = [
                d,
                group,
                code,
                name,
                _num(r[2]) if len(r) > 2 else None,   # 综合分
                _num(r[3]) if len(r) > 3 else None,   # 收盘
                _num(r[4]) if len(r) > 4 else None,   # 涨跌幅%
                r[5].strip() if len(r) > 5 else "",   # 状态
                r[6].strip() if len(r) > 6 else "",   # 提示
                _num(r[7]) if len(r) > 7 else None,   # MA10
                _num(r[8]) if len(r) > 8 else None,   # MA20
                _num(r[9]) if len(r) > 9 else None,   # 量比
                _num(r[10]) if len(r) > 10 else None,  # 10日涨幅%
                None,                                  # 买点序号（稍后填）
            ]
            rows.append(row)
            if d == latest:
                if group == "观察池" and row[7] in BUY_STATUSES:
                    chg10 = row[12]
                    if chg10 is None or chg10 < DEFAULT_MAX_10D_GAIN:
                        latest_buy_rows.append(len(rows) + 1)  # Data 第 1 行为表头
    watch_size = _watch_count(_parse_csv(csvs[-1][1]))
    return rows, len(csvs), latest_buy_rows, watch_size


def _active_position(state_dir: Path) -> dict | None:
    """读取纸面持仓，返回第一个 active 持仓（按买入日排序）。"""
    p = state_dir / "mainrise_positions.csv"
    if not p.exists():
        return None
    with open(p, encoding="utf-8-sig", newline="") as f:
        recs = list(csv.DictReader(f))
    acts = [r for r in recs if str(r.get("status", "")).strip() == "active"]
    if not acts:
        return None
    r = sorted(acts, key=lambda x: x.get("buy_date") or "")[0]
    try:
        price = float(r.get("buy_price") or 0)
    except (TypeError, ValueError):
        price = 0.0
    return {
        "code": str(r.get("code") or "601899").strip(),
        "name": str(r.get("name") or "紫金矿业").strip(),
        "buy_price": price if price > 0 else 32.12,
        "buy_date": str(r.get("buy_date") or "").strip(),
        "reason": str(r.get("reason") or "").strip(),
    }


# ---------------------------------------------------------------- 样式工具

def _style(ws, coord: str, proto: str):
    """把 proto 单元格的样式复制到 coord 单元格（返回该单元格）。"""
    cell = ws[coord]
    cell._style = copy(ws[proto]._style)
    return cell


def _put(ws, coord: str, value, proto: str | None = None):
    cell = ws[coord]
    if proto is not None:
        cell._style = copy(ws[proto]._style)
    cell.value = value
    return cell


# ---------------------------------------------------------------- 各表更新

def _fill_data(ws, rows: list[list], latest_buy_rows: list[int]):
    template_max = ws.max_row  # 模板 Data 现有行数（样式继承边界）
    for i, row in enumerate(rows, start=2):
        for c, val in enumerate(row, start=1):
            coord = f"{get_column_letter(c)}{i}"
            proto = f"{get_column_letter(c)}2" if i > template_max else None
            _put(ws, coord, val, proto)
    # 行数变少时清掉模板残留的旧行
    for r in range(len(rows) + 2, template_max + 1):
        for c in range(1, 15):
            ws.cell(r, c).value = None
    # 买点序号（静态行号，公式引擎不支持中文比较，故构建期写入）
    for r in range(2, len(rows) + 2):
        ws.cell(r, 14).value = r if r in latest_buy_rows else None
    # 超出模板的行补行高
    if len(rows) + 1 > template_max:
        h = ws.row_dimensions[2].height or 15
        for r in range(template_max + 1, len(rows) + 2):
            ws.row_dimensions[r].height = h


def _fill_watch(ws, csvs: list[tuple[date, Path]]):
    """最新日观察池（有综合分的行），按综合分降序；H/I 分档公式。"""
    watch_max = ws.max_row  # 模板 Watch 现有行数（样式继承边界）
    latest = csvs[-1]
    raw = _parse_csv(latest[1])
    items = []
    for r in raw:
        if not (len(r) > 2 and str(r[2]).strip()):
            continue  # 新信号行（综合分为空）不属于观察池
        items.append({
            "code": r[0].strip(),
            "name": r[1].strip(),
            "composite": _num(r[2]) if len(r) > 2 else None,
            "close": _num(r[3]) if len(r) > 3 else None,
            "chg": _num(r[4]) if len(r) > 4 else None,
            "status": r[5].strip() if len(r) > 5 else "",
            "hint": r[6].strip() if len(r) > 6 else "",
        })
    items.sort(key=lambda x: (-(x["composite"] or 0), x["code"]))
    for i, it in enumerate(items, start=2):
        for c, key in enumerate(("code", "name", "composite", "close", "chg", "status", "hint"), start=1):
            cell = ws.cell(i, c)
            if i > watch_max:
                cell._style = copy(ws.cell(2, c)._style)
            cell.value = it[key]
    for r in range(len(items) + 2, watch_max + 1):
        for c in range(1, 10):
            ws.cell(r, c).value = None
    for r in range(2, 12):
        ws.cell(r, 8).value = f'=IF($C{r}>80,$C{r},"")'
        ws.cell(r, 9).value = f'=IF($C{r}>75,IF($C{r}<=80,$C{r},""),"")'


def _fill_summary(ws, dates: list[date], pos: dict, watch_size: int) -> dict:
    """重写 Summary：按日表 + 状态分布 + 持仓收益；返回动态布局坐标。"""
    n = len(dates)
    tpl_max = ws.max_row
    per_day_last = n + 1
    cross_section = per_day_last + 1
    cross_header = cross_section + 1
    cross_last = cross_header + len(STATUSES)
    note_row = cross_last + 1
    pos_section = note_row + 2
    pos_header = pos_section + 1
    pos_start = pos_header + 1
    pos_last = pos_start + n - 1
    buy_label = pos_last + 2
    buy_value = buy_label + 1
    buy_note = buy_value + 1

    # 清空模板旧内容（行 1 表头保留；只清模板已有行，避免撑大维度）
    for r in range(2, tpl_max + 1):
        for c in range(1, 18):
            ws.cell(r, c).value = None

    # 按日聚合表（A..J 指标 + L 日期文本）
    for j, d in enumerate(dates):
        r = 2 + j
        _put(ws, f"A{r}", f"=DATE({d.year},{d.month},{d.day})", PER_DAY_PROTO["A"])
        _put(ws, f"B{r}",
             f'=COUNTIFS(Data!$A$2:$A${MAX_DATA_ROW},$A{r},Data!$B$2:$B${MAX_DATA_ROW},"观察池")',
             PER_DAY_PROTO["B"])
        _put(ws, f"C{r}",
             f'=IFERROR(AVERAGEIFS(Data!$E$2:$E${MAX_DATA_ROW},Data!$A$2:$A${MAX_DATA_ROW},$A{r},Data!$B$2:$B${MAX_DATA_ROW},"观察池"),0)',
             PER_DAY_PROTO["C"])
        _put(ws, f"D{r}",
             f'=IFERROR(AVERAGEIFS(Data!$G$2:$G${MAX_DATA_ROW},Data!$A$2:$A${MAX_DATA_ROW},$A{r},Data!$B$2:$B${MAX_DATA_ROW},"观察池"),0)',
             PER_DAY_PROTO["D"])
        _put(ws, f"E{r}",
             f'=COUNTIFS(Data!$A$2:$A${MAX_DATA_ROW},$A{r},Data!$B$2:$B${MAX_DATA_ROW},"观察池",Data!$G$2:$G${MAX_DATA_ROW},">0")',
             PER_DAY_PROTO["E"])
        _put(ws, f"F{r}",
             f'=COUNTIFS(Data!$A$2:$A${MAX_DATA_ROW},$A{r},Data!$B$2:$B${MAX_DATA_ROW},"观察池",Data!$H$2:$H${MAX_DATA_ROW},"T1确认买点",Data!$M$2:$M${MAX_DATA_ROW},"<{DEFAULT_MAX_10D_GAIN}")'
             f'+COUNTIFS(Data!$A$2:$A${MAX_DATA_ROW},$A{r},Data!$B$2:$B${MAX_DATA_ROW},"观察池",Data!$H$2:$H${MAX_DATA_ROW},"T0新信号",Data!$M$2:$M${MAX_DATA_ROW},"<{DEFAULT_MAX_10D_GAIN}")'
             f'+COUNTIFS(Data!$A$2:$A${MAX_DATA_ROW},$A{r},Data!$B$2:$B${MAX_DATA_ROW},"观察池",Data!$H$2:$H${MAX_DATA_ROW},"回踩低吸",Data!$M$2:$M${MAX_DATA_ROW},"<{DEFAULT_MAX_10D_GAIN}")',
             PER_DAY_PROTO["F"])
        _put(ws, f"G{r}",
             f'=COUNTIFS(Data!$A$2:$A${MAX_DATA_ROW},$A{r},Data!$B$2:$B${MAX_DATA_ROW},"观察池",Data!$H$2:$H${MAX_DATA_ROW},"破位")',
             PER_DAY_PROTO["G"])
        _put(ws, f"H{r}",
             f'=COUNTIFS(Data!$A$2:$A${MAX_DATA_ROW},$A{r},Data!$B$2:$B${MAX_DATA_ROW},"观察池",Data!$H$2:$H${MAX_DATA_ROW},"空头")',
             PER_DAY_PROTO["H"])
        _put(ws, f"I{r}",
             f'=COUNTIFS(Data!$A$2:$A${MAX_DATA_ROW},$A{r},Data!$B$2:$B${MAX_DATA_ROW},"观察池",Data!$H$2:$H${MAX_DATA_ROW},"多头持有")',
             PER_DAY_PROTO["I"])
        _put(ws, f"J{r}",
             f'=COUNTIFS(Data!$A$2:$A${MAX_DATA_ROW},$A{r},Data!$B$2:$B${MAX_DATA_ROW},"新信号")',
             PER_DAY_PROTO["J"])
        _put(ws, f"L{r}", f'=TEXT(A{r},"yyyy-mm-dd")', PER_DAY_PROTO["L"])

    # 状态分布（观察池，按日）
    _put(ws, f"A{cross_section}", "状态分布（观察池，按日）", CROSS_SECTION_PROTO)
    _put(ws, f"A{cross_header}", "状态", CROSS_HEADER_STATUS_PROTO)
    for j in range(n):
        col = get_column_letter(2 + j)
        _put(ws, f"{col}{cross_header}", f"=$A${2 + j}", CROSS_HEADER_DATE_PROTO)
    for k, status in enumerate(STATUSES, start=1):
        r = cross_header + k
        _put(ws, f"A{r}", status, CROSS_STATUS_LABEL_PROTO)
        for j in range(n):
            col = get_column_letter(2 + j)
            _put(ws, f"{col}{r}",
                 f'=COUNTIFS(Data!$A$2:$A${MAX_DATA_ROW},$A${2 + j},Data!$B$2:$B${MAX_DATA_ROW},"观察池",Data!$H$2:$H${MAX_DATA_ROW},$A{r})',
                 CROSS_STATUS_DATA_PROTO)

    # 数据源说明
    date_str = "/".join(d.isoformat() for d in dates)
    _put(ws, f"A{note_row}",
         f"数据源: output/reports/主升浪跟踪_{date_str}.csv（观察池{watch_size}只/日 + 全市场新信号按日计入）；"
         "涨跌幅/10日涨幅为百分数（8.83 即 8.83%）；"
         f"买点提示=观察池中 T0/T1/回踩低吸 且 10日涨幅<{DEFAULT_MAX_10D_GAIN:g}%（防追高）。",
         NOTE_PROTO)

    # 图表辅助数据：持仓收益
    _put(ws, f"A{pos_section}", "图表辅助数据（持仓收益 / Top10 分档）", POS_SECTION_PROTO)
    _put(ws, f"O{pos_header}", f"{pos['name']}收盘", POS_HEADER_PRICE_PROTO)
    _put(ws, f"P{pos_header}", "持仓盈亏%", POS_HEADER_PNL_PROTO)
    for j in range(n):
        r = pos_start + j
        _put(ws, f"O{r}",
             f'=SUMIFS(Data!$F$2:$F${MAX_DATA_ROW},Data!$A$2:$A${MAX_DATA_ROW},$A${2 + j},Data!$C$2:$C${MAX_DATA_ROW},"{pos["code"]}")',
             POS_DATA_PRICE_PROTO)
        _put(ws, f"P{r}", f'=IFERROR((O{r}/$P${buy_value}-1)*100,"")', POS_DATA_PNL_PROTO)
    _put(ws, f"P{buy_label}", f"买入价({pos['name']})", BUY_LABEL_PROTO)
    _put(ws, f"P{buy_value}", pos["buy_price"], BUY_VALUE_PROTO)
    _put(ws, f"P{buy_note}",
         f"来源: output/state/mainrise_positions.csv"
         + (f"（{pos['buy_date']} 建仓，当前 active）" if pos.get("buy_date") else ""),
         BUY_NOTE_PROTO)

    return {
        "per_day_last": per_day_last,
        "cross_header": cross_header,
        "note_row": note_row,
        "pos_header": pos_header,
        "pos_start": pos_start,
        "pos_last": pos_last,
        "buy_value": buy_value,
    }


def _fill_dashboard(ws, n: int, pos: dict, buy_count: int, layout: dict, dates: list[date]):
    r = layout["per_day_last"]
    # KPI 卡片（最新日）
    ws["B3"] = f"=COUNTA(Summary!A2:A{r})"
    ws["D3"] = f"=Summary!B{r}"
    ws["F3"] = f"=Summary!F{r}"
    ws["H3"] = f"=Summary!C{r}"
    ws["J3"] = f"=Summary!D{r}"
    ws["L3"] = f"=Summary!E{r}"
    # 今日买点提示标题
    ws["B7"] = f'="今日买点提示（"&TEXT(Summary!$A${r},"yyyy-mm-dd")&"）"'
    # 买点明细行（模板固定 3 行，超出时复制样式/合并扩展）
    rows_needed = max(3, buy_count)
    for k in range(1, rows_needed + 1):
        rr = 8 + k
        if rr > 11:
            for c in ("B", "C", "D", "E", "F", "G", "H", "I"):
                _style(ws, f"{c}{rr}", f"{c}11")
            for m in (("C", "D"), ("E", "F"), ("G", "I")):
                ws.merge_cells(f"{m[0]}{rr}:{m[1]}{rr}")
            if ws.row_dimensions[11].height:
                ws.row_dimensions[rr].height = ws.row_dimensions[11].height
        ws[f"B{rr}"] = f'=IFERROR(INDEX(Data!$H:$H,SMALL(Data!$N$2:$N${MAX_DATA_ROW},{k})),"")'
        ws[f"C{rr}"] = f'=IFERROR(INDEX(Data!$D:$D,SMALL(Data!$N$2:$N${MAX_DATA_ROW},{k})),"")'
        ws[f"E{rr}"] = f'=IFERROR(INDEX(Data!$E:$E,SMALL(Data!$N$2:$N${MAX_DATA_ROW},{k})),"")'
        ws[f"G{rr}"] = f'=IFERROR(INDEX(Data!$I:$I,SMALL(Data!$N$2:$N${MAX_DATA_ROW},{k})),"")'
    # 脚注
    date_str = "/".join(d.isoformat() for d in dates)
    ws["B32"] = (f"主升浪信号跟踪 KPI 仪表盘（数据源: output/reports/主升浪跟踪_{date_str}.csv"
                 " + output/state/mainrise_positions.csv；涨跌幅/10日涨幅为百分数，如 8.83 即 8.83%）")
    ws["B33"] = ("买点明细=最新日观察池内 T0/T1/回踩低吸 且 "
                 f"10日涨幅<{DEFAULT_MAX_10D_GAIN:g}%（防追高）；"
                 f"持仓收益=每日收盘/买入价-1（买入价 {pos['buy_price']:g}，来源 positions.csv）；"
                 "免责：研究线索，不构成投资建议")


# ---------------------------------------------------------------- 图表 XML 补丁

def _patch_chart_xml(xml: bytes, n: int, pos_start: int, pos_name: str | None) -> bytes:
    text = xml.decode("utf-8")
    text = re.sub(r'(Summary!\$D\$2:\$D\$)\d+', rf"\g<1>{n + 1}", text)
    text = re.sub(r'(Summary!\$L\$2:\$L\$)\d+', rf"\g<1>{n + 1}", text)
    text = re.sub(r'(Summary!\$P\$)\d+(:\$P\$)\d+', rf"\g<1>{pos_start}\g<2>{pos_start + n - 1}", text)
    text = re.sub(r"观察池平均涨跌幅（三日）", f"观察池平均涨跌幅（{n}日）", text)
    if pos_name and "持仓收益曲线" in text:
        text = text.replace("紫金矿业", pos_name)
    return text.encode("utf-8")


def _ordered_chart_parts(zf: zipfile.ZipFile) -> list[str]:
    """按 drawing1.xml 中图表出现顺序返回 chart 部件完整路径列表（如 xl/charts/chart1.xml）。"""
    rels: dict[str, str] = {}
    try:
        rels_xml = zf.read("xl/drawings/_rels/drawing1.xml.rels").decode("utf-8")
        for rel in re.finditer(r"<Relationship[^>]*>", rels_xml):
            attrs = rel.group(0)
            rid = re.search(r'Id="(rId\d+)"', attrs)
            target = re.search(r'Target="([^"]+\.xml)"', attrs)
            if rid and target:
                t = target.group(1)
                if t.startswith("/"):
                    rels[rid.group(1)] = t.lstrip("/")
                elif t.startswith("../"):
                    rels[rid.group(1)] = "xl/" + t[3:]
                else:
                    rels[rid.group(1)] = "xl/drawings/" + t
        drawing = zf.read("xl/drawings/drawing1.xml").decode("utf-8")
        ids = re.findall(r'<c:chart[^>]*r:id="(rId\d+)"', drawing)
    except KeyError:
        return []
    return [rels[i] for i in ids if rels.get(i)]


def _restore_chart_parts(output: Path, template: Path, n: int, pos_start: int, pos_name: str | None):
    """用模板原始 chart XML（仅改数据引用）覆盖 openpyxl 重序列化后的图表，保住样式。"""
    with zipfile.ZipFile(template) as zt, zipfile.ZipFile(output) as zo:
        tpl_parts = _ordered_chart_parts(zt)
        out_parts = _ordered_chart_parts(zo)
        if not tpl_parts or len(tpl_parts) != len(out_parts):
            return  # 结构异常时保留 openpyxl 输出，不冒险覆盖
        patched = {}
        for t_name, o_name in zip(tpl_parts, out_parts):
            xml = zt.read(t_name)
            patched[o_name] = _patch_chart_xml(xml, n, pos_start, pos_name)
    tmp = output.with_suffix(".tmp.xlsx")
    with zipfile.ZipFile(output) as zin, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename in patched:
                data = patched[item.filename]
            zout.writestr(item, data)
    tmp.replace(output)


def _fix_styles_fonts(output: Path):
    """openpyxl 重写 styles.xml 后 font/dxf-font 的子元素顺序不合 schema，
    按模板约定顺序（b, sz, color, name, family ...）重排，避免严格校验器报错。"""
    ns = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    order = ["b", "i", "strike", "sz", "color", "u", "name", "family", "charset", "scheme"]
    with zipfile.ZipFile(output) as z:
        styles = z.read("xl/styles.xml")
    root = ET.fromstring(styles)
    font_nodes = list(root.findall("m:fonts/m:font", ns))
    for dxf in root.findall("m:dxfs/m:dxf", ns):
        f = dxf.find("m:font", ns)
        if f is not None:
            font_nodes.append(f)
    for font in font_nodes:
        if font is None or not len(font):
            continue
        by_tag: dict[str, list] = {}
        for child in list(font):
            by_tag.setdefault(child.tag.rsplit("}", 1)[-1], []).append(child)
        ordered: list = []
        for tag in order:
            ordered.extend(by_tag.pop(tag, []))
        for rest in by_tag.values():
            ordered.extend(rest)
        for child in list(font):
            font.remove(child)
        for child in ordered:
            font.append(child)
    ET.register_namespace("", ns["m"])
    new_styles = ET.tostring(root, encoding="utf-8", xml_declaration=True)
    tmp = output.with_suffix(".styles.tmp.xlsx")
    with zipfile.ZipFile(output) as zin, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == "xl/styles.xml":
                data = new_styles
            zout.writestr(item, data)
    tmp.replace(output)


# ---------------------------------------------------------------- 入口

def _resolve_template(explicit: Path | None, reports_dir: Path) -> Path:
    if explicit is not None:
        return explicit
    local = reports_dir / QUANTDARK_TEMPLATE_NAME
    if local.exists():
        return local
    # 内置模板：源码模式用 __file__ 相对路径；PyInstaller 冻结模式用 sys._MEIPASS
    candidates = [BUNDLED_TEMPLATE]
    meipass = getattr(sys, "_MEIPASS", None)
    if meipass:
        candidates.append(Path(meipass) / "mainrise" / "resources" / "dashboard_template.xlsx")
    for cand in candidates:
        if cand.exists():
            return cand
    raise FileNotFoundError(
        f"未找到仪表盘模板（{QUANTDARK_TEMPLATE_NAME} 或内置 dashboard_template.xlsx）")


def update_dashboard(template: str | Path | None = None,
                     output: str | Path | None = None,
                     reports_dir: str | Path | None = None,
                     state_dir: str | Path | None = None) -> Path:
    """以模板为基础，用最新跟踪数据重新生成仪表盘，返回输出路径。"""
    reports = Path(reports_dir) if reports_dir else paths.report_dir()
    state = Path(state_dir) if state_dir else paths.state_dir()
    tpl = _resolve_template(Path(template) if template else None, reports)
    out = Path(output) if output else reports / OUTPUT_NAME

    csvs = _tracking_csvs(reports)
    if not csvs:
        raise ValueError(f"{reports} 下没有跟踪 CSV（{TRACK_PREFIX}*.csv），请先运行 mainrise track")

    rows, n_days, buy_rows, watch_size = _data_rows(csvs)
    dates = [d for d, _ in csvs]
    pos = _active_position(state) or {
        "code": "601899", "name": "紫金矿业", "buy_price": 32.12,
        "buy_date": "2026-08-04", "reason": ""}

    out.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(tpl, out)

    wb = load_workbook(out)
    _fill_data(wb["Data"], rows, buy_rows)
    _fill_watch(wb["Watch"], csvs)
    layout = _fill_summary(wb["Summary"], dates, pos, watch_size)
    _fill_dashboard(wb["Dashboard"], n_days, pos, len(buy_rows), layout, dates)
    wb.calculation.fullCalcOnLoad = True
    wb.save(out)

    _restore_chart_parts(out, tpl, n_days, layout["pos_start"], pos["name"])
    _fix_styles_fonts(out)
    return out
