"""KPI 仪表盘自动更新测试：模板为基础重生成，数据/公式/图表引用随天数伸缩。"""
import csv
import tempfile
import unittest
import zipfile
from pathlib import Path

from openpyxl import load_workbook

from mainrise.dashboard import update_dashboard


def _watch_csv(day: str, closes: list) -> list:
    header = ["code", "name", "composite", "close", "chg", "status", "hint",
              "ma10", "ma20", "vr", "chg10"]
    rows = [header]
    base = [("000001", "平安银行", 86.3), ("000002", "万科A", 81.8),
            ("000003", "测试股三", 78.0), ("000004", "测试股四", 74.5)]
    for i, (code, name, comp) in enumerate(base * 9):
        if i >= 36:
            break
        close = closes[i % len(closes)]
        status, hint, chg10 = "空头", "均线未多头，观望", 2.0
        if code == "000002":
            status, hint, chg10 = "二波加仓", "最优买点：明日开盘加仓", 5.0
        elif code == "000003":
            status, hint, chg10 = "B3打底仓", "均线粘合爆量突破：明日打底仓", 10.0
        rows.append([code, name, comp, close, 1.0, status, hint,
                     10.0, 9.5, 1.2, chg10])
    return rows


def _positions_csv() -> list:
    return [["code", "name", "signal_date", "confirm_date", "buy_kind",
             "buy_date", "buy_price", "peak", "peak_date", "status",
             "close_date", "close_price", "reason"],
            ["000002", "万科A", "", "", 2, "2026-08-04", 10.0,
             11.0, "2026-08-05", "active", "", "", "B3打底仓"]]


def _write(reports: Path, states: Path) -> None:
    for day, closes in (("2026-08-04", [10.0, 10.2]),
                        ("2026-08-05", [10.4, 10.5])):
        with open(reports / f"主升浪跟踪_{day}.csv", "w", newline="",
                  encoding="utf-8") as f:
            csv.writer(f).writerows(_watch_csv(day, closes))
    with open(states / "mainrise_positions.csv", "w", newline="",
              encoding="utf-8") as f:
        csv.writer(f).writerows(_positions_csv())


class TestDashboard(unittest.TestCase):
    def _run(self, reports: Path, states: Path) -> Path:
        return update_dashboard(reports_dir=reports, state_dir=states)

    def test_basic_update(self):
        with tempfile.TemporaryDirectory() as td:
            td = Path(td)
            reports, states = td / "reports", td / "state"
            reports.mkdir(), states.mkdir()
            _write(reports, states)
            out = self._run(reports, states)
            self.assertTrue(out.exists())
            wb = load_workbook(out)
            dash = wb["Dashboard"]
            self.assertEqual(dash["F3"].value, "=Summary!F3")  # 最新日=第2行数据
            self.assertEqual(dash["B7"].value,
                             '="今日买点提示（"&TEXT(Summary!$A$3,"yyyy-mm-dd")&"）"')
            data = wb["Data"]
            # 第 2 日(最新)36 行观察池：000002 在第 2 行 -> Data 39，000003 -> 40
            self.assertEqual(data["N39"].value, 39)
            self.assertEqual(data["N40"].value, 40)
            self.assertEqual(data["N38"].value, None)
            summ = wb["Summary"]
            self.assertEqual(summ["A2"].value, "=DATE(2026,8,4)")
            self.assertEqual(summ["A3"].value, "=DATE(2026,8,5)")
            self.assertTrue(wb.calculation.fullCalcOnLoad)
            # 图表引用随天数伸缩（n=2：持仓 P 区间起始行=17）
            with zipfile.ZipFile(out) as z:
                chart3 = z.read("xl/charts/chart3.xml").decode("utf-8")
            self.assertIn("Summary!$P$14:$P$15", chart3)
            self.assertIn("Summary!$L$2:$L$3", chart3)

    def test_growth_updates_latest_row(self):
        with tempfile.TemporaryDirectory() as td:
            td = Path(td)
            reports, states = td / "reports", td / "state"
            reports.mkdir(), states.mkdir()
            _write(reports, states)
            # 追加第 3 天
            with open(reports / "主升浪跟踪_2026-08-06.csv", "w",
                      newline="", encoding="utf-8") as f:
                csv.writer(f).writerows(_watch_csv("2026-08-06", [10.6, 10.7]))
            out = self._run(reports, states)
            wb = load_workbook(out)
            dash = wb["Dashboard"]
            self.assertEqual(dash["F3"].value, "=Summary!F4")  # 最新日=第3行
            self.assertEqual(dash["B3"].value, "=COUNTA(Summary!A2:A4)")
            data = wb["Data"]
            # 第 3 日(最新)起始行 74：000002 -> 75，000003 -> 76
            self.assertEqual(data["N75"].value, 75)
            self.assertEqual(data["N76"].value, 76)
            with zipfile.ZipFile(out) as z:
                chart1 = z.read("xl/charts/chart1.xml").decode("utf-8")
                chart3 = z.read("xl/charts/chart3.xml").decode("utf-8")
            self.assertIn("Summary!$D$2:$D$4", chart1)
            self.assertIn("Summary!$P$15:$P$17", chart3)

    def test_dynamic_watchlist_size(self):
        """观察池数量不再写死 36：按“有综合分的连续行”自动判定。"""
        with tempfile.TemporaryDirectory() as td:
            td = Path(td)
            reports, states = td / "reports", td / "state"
            reports.mkdir(), states.mkdir()
            header = ["code", "name", "composite", "close", "chg", "status",
                      "hint", "ma10", "ma20", "vr", "chg10"]
            rows = [header]
            for i in range(1, 41):  # 40 只观察池
                status, hint, chg10 = "空头", "均线未多头，观望", 2.0
                if i == 5:
                    status, hint, chg10 = "二波加仓", "最优买点：明日开盘加仓", 5.0
                rows.append([f"0000{i:02d}", f"股{i}", 70.0 + i, 10.0, 1.0,
                             status, hint, 10.0, 9.5, 1.2, chg10])
            for i in range(41, 44):  # 3 只新信号（综合分为空）
                rows.append([f"0000{i:02d}", f"新{i}", "", "", 5.0,
                             "B3打底仓", "新信号，待财务评估后入池", "", "", 2.0, 8.0])
            with open(reports / "主升浪跟踪_2026-08-04.csv", "w", newline="",
                      encoding="utf-8") as f:
                csv.writer(f).writerows(rows)
            out = self._run(reports, states)
            wb = load_workbook(out)
            data = wb["Data"]
            self.assertEqual(data["B2"].value, "观察池")
            self.assertEqual(data["B41"].value, "观察池")
            self.assertEqual(data["B42"].value, "新信号")
            self.assertEqual(data["N6"].value, 6)  # 第 5 只（二波加仓）-> Data 行 6
            summ = wb["Summary"]
            self.assertIn("观察池40只/日", summ["A9"].value)  # n=1 时脚注在 A9
            self.assertIn('"新信号"', summ["J2"].value)
            watch = wb["Watch"]
            self.assertEqual(watch["A2"].value, "000040")   # 综合分最高排最前
            self.assertEqual(watch["A41"].value, "000001")  # 40 只全部写入


if __name__ == "__main__":
    unittest.main()
